import os
from datetime import datetime
from database.db import get_connection
from core.config import FILE_LOCK, EXCEL_FILE, REQUIRED_COLUMNS, logger

def get_all_entries() -> list[dict]:
    with FILE_LOCK:
        conn = get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, date AS "تاریخ", course AS "درس", topic AS "مبحث",
                       duration AS "زمان (دقیقه)", tests AS "تعداد تست",
                       wrongs AS "تست‌های مرور", notes AS "یادداشت"
                FROM study_logs
                ORDER BY id DESC
            """)
            rows = cursor.fetchall()
            return [dict(row) for row in rows]
        finally:
            conn.close()

def add_entry(course: str, topic: str, duration: int, tests: int, wrongs: str, notes: str) -> int:
    with FILE_LOCK:
        duration = max(0, duration)
        tests = max(0, tests)
        date_str = datetime.now().strftime("%Y-%m-%d")
        course_clean = course.strip() if course else "سایر"
        topic_clean = topic.strip() if topic and topic.strip() else "-"
        wrongs_clean = wrongs.strip() if wrongs and wrongs.strip() else "-"
        notes_clean = notes.strip() if notes and notes.strip() else "-"

        conn = get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO study_logs (date, course, topic, duration, tests, wrongs, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (date_str, course_clean, topic_clean, duration, tests, wrongs_clean, notes_clean))
            conn.commit()
            new_id = cursor.lastrowid
            logger.info(f"Added new study entry ID={new_id}")
            return new_id
        finally:
            conn.close()

def delete_entry(item_id: int) -> bool:
    with FILE_LOCK:
        conn = get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM study_logs WHERE id = ?", (item_id,))
            conn.commit()
            logger.info(f"Deleted study entry ID={item_id}")
            return cursor.rowcount > 0
        finally:
            conn.close()

def get_today_stats() -> tuple[float, int]:
    today_str = datetime.now().strftime("%Y-%m-%d")
    with FILE_LOCK:
        conn = get_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT SUM(duration) as total_duration, SUM(tests) as total_tests
                FROM study_logs
                WHERE date = ?
            """, (today_str,))
            row = cursor.fetchone()
            total_minutes = row["total_duration"] or 0
            total_tests = row["total_tests"] or 0
            today_hours = round(float(total_minutes) / 60.0, 1)
            return today_hours, int(total_tests)
        finally:
            conn.close()

def export_to_excel() -> str:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    with FILE_LOCK:
        conn = get_connection()
        try:
            df = pd.read_sql_query("""
                SELECT id AS "شناسه", date AS "تاریخ", course AS "درس", topic AS "مبحث",
                       duration AS "زمان (دقیقه)", tests AS "تعداد تست",
                       wrongs AS "تست‌های مرور", notes AS "یادداشت"
                FROM study_logs
                ORDER BY id ASC
            """, conn)
            
            headers = ["شناسه", "تاریخ", "درس", "مبحث", "زمان (دقیقه)", "تعداد تست", "تست‌های مرور", "یادداشت"]
            if df.empty:
                df = pd.DataFrame(columns=headers)
                
            temp_excel = EXCEL_FILE + ".tmp.xlsx"
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "سوابق مطالعه"
            
            # 1. Right-to-Left Sheet Layout
            ws.sheet_view.rightToLeft = True

            # Fills (App Indigo & Emerald Theme)
            brand_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
            btn_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
            header_fill = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
            zebra_even_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            summary_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")

            # Fonts
            brand_title_font = Font(name="Vazirmatn", size=14, bold=True, color="3730A3")
            brand_sub_font = Font(name="Vazirmatn", size=9.5, color="6366F1")
            brand_meta_font = Font(name="Vazirmatn", size=9.5, bold=True, color="334155")
            btn_font = Font(name="Vazirmatn", size=9.5, bold=True, color="FFFFFF")
            header_font = Font(name="Vazirmatn", size=10.5, bold=True, color="FFFFFF")
            data_font = Font(name="Vazirmatn", size=10, color="0F172A")
            summary_font = Font(name="Vazirmatn", size=10.5, bold=True, color="047857")

            align_center = Alignment(horizontal="center", vertical="center")
            align_center_rtl = Alignment(horizontal="center", vertical="center", readingOrder=2)
            align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
            
            thin_border = Border(
                left=Side(style='thin', color='E2E8F0'),
                right=Side(style='thin', color='E2E8F0'),
                top=Side(style='thin', color='E2E8F0'),
                bottom=Side(style='thin', color='E2E8F0')
            )
            summary_border = Border(
                top=Side(style='medium', color='10B981'),
                bottom=Side(style='double', color='10B981'),
                left=Side(style='thin', color='E2E8F0'),
                right=Side(style='thin', color='E2E8F0')
            )

            # ROW 1: Brand Title Banner (A1:H1)
            ws.merge_cells("A1:H1")
            cell_title = ws["A1"]
            cell_title.value = "⏰ چند ساعت؟ — نرم‌افزار ثبت ساعات مطالعه و تمرکز"
            cell_title.font = brand_title_font
            cell_title.fill = brand_fill
            cell_title.alignment = align_center_rtl

            for c in range(1, 9):
                ws.cell(row=1, column=c).fill = brand_fill

            ws.row_dimensions[1].height = 42

            # ROW 2: Subtitle
            ws.merge_cells("A2:H2")
            cell_sub = ws["A2"]
            cell_sub.value = "برنامه‌ریزی هوشمند، ثبت پارت‌های پومودورو و تحلیل پیشرفت تحصیلی"
            cell_sub.font = brand_sub_font
            cell_sub.fill = brand_fill
            cell_sub.alignment = align_center_rtl
            ws.row_dimensions[2].height = 22

            # ROW 3: Export Metadata (A3:E3) + Green Download Button Cell (F3:H3) (Both explicit RTL reading order)
            today_str = datetime.now().strftime("%Y-%m-%d")
            ws.merge_cells("A3:E3")
            cell_meta = ws["A3"]
            cell_meta.value = f"📅 تاریخ خروجی: {today_str}    |    📊 تعداد کل جلسات ثبت‌شده: {len(df)} جلسه"
            cell_meta.font = brand_meta_font
            cell_meta.alignment = align_center_rtl

            ws.merge_cells("F3:H3")
            cell_btn = ws["F3"]
            cell_btn.value = "📥 وب‌سایت رسمی و دانلود آخرین نسخه"
            cell_btn.hyperlink = "https://yahaiz.github.io/chand-saat/"
            cell_btn.font = btn_font
            cell_btn.fill = btn_fill
            cell_btn.alignment = align_center_rtl
            ws.row_dimensions[3].height = 26

            # ROW 4: Separator Gap
            ws.row_dimensions[4].height = 10

            # ROW 5: Table Header
            table_start_row = 5
            ws.row_dimensions[table_start_row].height = 26
            for col_idx, h_text in enumerate(headers, 1):
                cell = ws.cell(row=table_start_row, column=col_idx, value=h_text)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center

            # ROW 6+: Data Rows
            data_start_row = 6
            for r_offset, row in enumerate(df.values):
                r_idx = data_start_row + r_offset
                ws.row_dimensions[r_idx].height = 22
                for c_idx, val in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.font = data_font
                    cell.border = thin_border
                    
                    if r_offset % 2 == 1:
                        cell.fill = zebra_even_fill
                        
                    if c_idx == 8:
                        cell.alignment = align_right
                    else:
                        cell.alignment = align_center

            data_end_row = data_start_row + len(df) - 1 if len(df) > 0 else table_start_row

            # ROW Summary (Total)
            summary_row = data_end_row + 1 if len(df) > 0 else 6
            ws.row_dimensions[summary_row].height = 26
            
            total_duration = int(df["زمان (دقیقه)"].sum()) if not df.empty else 0
            total_tests = int(df["تعداد تست"].sum()) if not df.empty else 0
            total_hours = round(total_duration / 60.0, 1)

            ws.cell(row=summary_row, column=1, value="مجموع").alignment = align_center
            ws.cell(row=summary_row, column=5, value=f"{total_duration} دقیقه ({total_hours} ساعت)").alignment = align_center
            ws.cell(row=summary_row, column=6, value=f"{total_tests} تست").alignment = align_center

            for c_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=summary_row, column=c_idx)
                cell.font = summary_font
                cell.fill = summary_fill
                cell.border = summary_border

            # Set Auto Filter on Table Range
            ws.auto_filter.ref = f"A{table_start_row}:H{data_end_row}"

            # Custom & Auto Column Widths
            for col in ws.columns:
                col_idx = col[0].column
                col_letter = get_column_letter(col_idx)
                
                # Custom requested column widths:
                # Column 5: "زمان (دقیقه)" -> reduced width (1/3 of default)
                if col_idx == 5:
                    ws.column_dimensions[col_letter].width = 15
                # Column 6: "تعداد تست" -> reduced width (1/2 of default)
                elif col_idx == 6:
                    ws.column_dimensions[col_letter].width = 12
                else:
                    max_len = 0
                    for cell in col:
                        if cell.row < table_start_row:
                            continue
                        val_str = str(cell.value or '')
                        length = sum(2 if ord(c) > 127 else 1 for c in val_str)
                        if length > max_len:
                            max_len = length
                    ws.column_dimensions[col_letter].width = max(max_len + 6, 14)

            wb.save(temp_excel)

            if os.path.exists(EXCEL_FILE):
                os.remove(EXCEL_FILE)
            os.rename(temp_excel, EXCEL_FILE)
            return EXCEL_FILE
        except Exception as e:
            logger.error(f"Error exporting database to Excel: {e}")
            df_empty = pd.DataFrame(columns=REQUIRED_COLUMNS)
            df_empty.to_excel(EXCEL_FILE, index=False)
            return EXCEL_FILE
        finally:
            conn.close()
